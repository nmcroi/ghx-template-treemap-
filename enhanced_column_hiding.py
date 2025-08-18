#!/usr/bin/env python3
"""
Enhanced Column Hiding Module voor GHX Template Generator

Deze module bevat een herbruikbare hide_columns() functie die verschillende
methodes test om kolommen permanent te verbergen in Excel bestanden.
"""

from pathlib import Path
from typing import List, Dict, Any, Optional
from enum import Enum
import json
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill


class HideMethod(Enum):
    """Verschillende methodes om kolommen te verbergen."""
    HIDDEN_ONLY = "hidden_only"
    WIDTH_ZERO = "width_zero"
    COMBINED = "combined"
    CONTENT_CLEAR = "content_clear"
    OUTLINE_COLLAPSE = "outline_collapse"
    ALL_METHODS = "all_methods"


class ColumnHider:
    """
    Klasse voor het verbergen van Excel kolommen met verschillende methodes.
    """
    
    def __init__(self):
        self.test_results = {}
    
    def hide_columns(self, 
                    worksheet: Worksheet, 
                    columns_to_hide: List[str],
                    method: HideMethod = HideMethod.ALL_METHODS,
                    clear_content: bool = False) -> Dict[str, Any]:
        """
        Verberg kolommen met gespecificeerde methode.
        
        Args:
            worksheet: Excel worksheet object
            columns_to_hide: Lijst van kolom letters (bijv. ['AA', 'AB'])
            method: Methode om te gebruiken voor verbergen
            clear_content: Of inhoud van kolommen gewist moet worden
            
        Returns:
            Dictionary met resultaten per kolom
        """
        results = {
            'method': method.value,
            'columns_processed': [],
            'errors': [],
            'success_count': 0
        }
        
        for col in columns_to_hide:
            try:
                col_result = self._apply_hide_method(worksheet, col, method, clear_content)
                results['columns_processed'].append({
                    'column': col,
                    'success': col_result['success'],
                    'actions': col_result['actions'],
                    'header': col_result['header']
                })
                
                if col_result['success']:
                    results['success_count'] += 1
                    
            except Exception as e:
                error_msg = f"Kolom {col}: {str(e)}"
                results['errors'].append(error_msg)
        
        return results
    
    def _apply_hide_method(self, 
                          worksheet: Worksheet, 
                          column: str, 
                          method: HideMethod,
                          clear_content: bool) -> Dict[str, Any]:
        """
        Pas specifieke verstop-methode toe op één kolom.
        
        Args:
            worksheet: Excel worksheet
            column: Kolom letter
            method: Verstop methode
            clear_content: Of inhoud gewist moet worden
            
        Returns:
            Resultaat dictionary
        """
        actions_taken = []
        success = True
        
        # Haal header tekst op voor logging
        try:
            header_cell = worksheet[f"{column}1"]
            header_text = header_cell.value or "Geen header"
        except:
            header_text = "Onbekend"
        
        try:
            if method == HideMethod.HIDDEN_ONLY:
                worksheet.column_dimensions[column].hidden = True
                actions_taken.append("hidden = True")
                
            elif method == HideMethod.WIDTH_ZERO:
                worksheet.column_dimensions[column].width = 0
                actions_taken.append("width = 0")
                
            elif method == HideMethod.COMBINED:
                worksheet.column_dimensions[column].hidden = True
                worksheet.column_dimensions[column].width = 0
                actions_taken.extend(["hidden = True", "width = 0"])
                
            elif method == HideMethod.OUTLINE_COLLAPSE:
                worksheet.column_dimensions[column].hidden = True
                worksheet.column_dimensions[column].outline_level = 1
                worksheet.column_dimensions[column].collapsed = True
                actions_taken.extend(["hidden = True", "outline_level = 1", "collapsed = True"])
                
            elif method == HideMethod.ALL_METHODS:
                # Combineer alle methodes voor maximale compatibiliteit
                worksheet.column_dimensions[column].hidden = True
                worksheet.column_dimensions[column].width = 0
                worksheet.column_dimensions[column].outline_level = 1
                try:
                    worksheet.column_dimensions[column].collapsed = True
                    actions_taken.append("collapsed = True")
                except:
                    pass  # Collapsed werkt niet altijd
                actions_taken.extend(["hidden = True", "width = 0", "outline_level = 1"])
            
            # Content clearing (optioneel)
            if clear_content or method == HideMethod.CONTENT_CLEAR:
                cleared_cells = self._clear_column_content(worksheet, column)
                actions_taken.append(f"cleared {cleared_cells} cells")
                
        except Exception as e:
            success = False
            actions_taken.append(f"ERROR: {str(e)}")
        
        return {
            'success': success,
            'actions': actions_taken,
            'header': header_text
        }
    
    def _clear_column_content(self, worksheet: Worksheet, column: str) -> int:
        """
        Wis inhoud van alle cellen in een kolom.
        
        Args:
            worksheet: Excel worksheet
            column: Kolom letter
            
        Returns:
            Aantal gewiste cellen
        """
        cleared_count = 0
        max_row = min(worksheet.max_row, 1000)  # Limiteer voor performance
        
        for row in range(1, max_row + 1):
            try:
                cell = worksheet[f"{column}{row}"]
                if cell.value is not None:
                    cell.value = None
                    cleared_count += 1
            except:
                continue
                
        return cleared_count
    
    def test_all_methods(self, 
                        input_file: Path, 
                        columns_to_hide: List[str] = ['AA', 'AB'],
                        sheet_name: str = "Template NL") -> Dict[str, Any]:
        """
        Test alle verstop-methodes op een bestand.
        
        Args:
            input_file: Pad naar input Excel bestand
            columns_to_hide: Kolommen om te verbergen
            sheet_name: Naam van worksheet
            
        Returns:
            Resultaten van alle tests
        """
        test_results = {
            'input_file': str(input_file),
            'columns_tested': columns_to_hide,
            'sheet_name': sheet_name,
            'methods_tested': {},
            'recommendations': []
        }
        
        # Test elke methode
        for method in HideMethod:
            print(f"Testing methode: {method.value}")
            
            try:
                # Laad fresh copy voor elke test
                wb = load_workbook(input_file)
                
                if sheet_name not in wb.sheetnames:
                    test_results['methods_tested'][method.value] = {
                        'error': f"Sheet '{sheet_name}' niet gevonden"
                    }
                    continue
                
                ws = wb[sheet_name]
                
                # Pas methode toe
                result = self.hide_columns(ws, columns_to_hide, method)
                
                # Sla test bestand op
                output_file = Path(f"out/test_hide_{method.value}.xlsx")
                output_file.parent.mkdir(exist_ok=True)
                wb.save(output_file)
                
                result['output_file'] = str(output_file)
                test_results['methods_tested'][method.value] = result
                
                print(f"  ✅ {method.value}: {result['success_count']}/{len(columns_to_hide)} kolommen verborgen")
                
            except Exception as e:
                test_results['methods_tested'][method.value] = {
                    'error': str(e)
                }
                print(f"  ❌ {method.value}: {str(e)}")
        
        # Genereer aanbevelingen
        test_results['recommendations'] = self._generate_recommendations(test_results)
        
        return test_results
    
    def _generate_recommendations(self, test_results: Dict[str, Any]) -> List[str]:
        """Genereer aanbevelingen gebaseerd op test resultaten."""
        recommendations = []
        
        # Analyseer succespercentages
        success_rates = {}
        for method_name, result in test_results['methods_tested'].items():
            if 'error' not in result:
                total_cols = len(test_results['columns_tested'])
                success_rate = result['success_count'] / total_cols if total_cols > 0 else 0
                success_rates[method_name] = success_rate
        
        if success_rates:
            best_method = max(success_rates.items(), key=lambda x: x[1])
            recommendations.append(f"Beste methode: {best_method[0]} ({best_method[1]*100:.0f}% succes)")
            
            if best_method[1] == 1.0:
                recommendations.append("✅ Perfecte compatibiliteit gevonden")
            elif best_method[1] >= 0.8:
                recommendations.append("⚠️ Goede maar niet perfecte compatibiliteit")
            else:
                recommendations.append("❌ Lage compatibiliteit - mogelijk problemen")
        
        # Specifieke aanbevelingen
        if 'all_methods' in success_rates and success_rates['all_methods'] >= 0.8:
            recommendations.append("Aanbeveling: Gebruik ALL_METHODS voor maximale compatibiliteit")
        elif 'combined' in success_rates and success_rates['combined'] >= 0.8:
            recommendations.append("Aanbeveling: Gebruik COMBINED methode")
        else:
            recommendations.append("Aanbeveling: Test handmatig in doeltoepassing")
            
        return recommendations


def integrate_with_excel_engine(excel_processor_path: Path = None):
    """
    Integratie met bestaande ExcelProcessor klasse.
    
    Args:
        excel_processor_path: Pad naar excel.py bestand
    """
    if not excel_processor_path:
        excel_processor_path = Path("src/excel.py")
    
    print(f"Integratie suggestie voor {excel_processor_path}:")
    print("-" * 50)
    
    integration_code = '''
    # Voeg toe aan ExcelProcessor klasse:
    
    def hide_columns_permanently(self, 
                                worksheet: Worksheet,
                                columns_to_hide: List[str],
                                method: str = "all_methods") -> None:
        """
        Verberg kolommen permanent met verschillende methodes.
        
        Args:
            worksheet: Excel worksheet
            columns_to_hide: Lijst van kolom letters
            method: Verstop methode ('all_methods' aanbevolen)
        """
        from enhanced_column_hiding import ColumnHider, HideMethod
        
        hider = ColumnHider()
        method_enum = HideMethod(method)
        
        result = hider.hide_columns(worksheet, columns_to_hide, method_enum)
        
        if result['errors']:
            print(f"Waarschuwingen bij kolom verbergen: {result['errors']}")
    
    # Gebruik in _apply_column_decisions methode:
    
    def _apply_column_decisions(self, ws: Worksheet, decisions_by_col: Dict[str, FieldDecision]) -> None:
        """Bestaande methode - voeg toe aan het eind:"""
        
        # ... bestaande code ...
        
        # Forceer verbergen van specifieke kolommen (AA, AB)
        force_hide_columns = ['AA', 'AB']
        if any(col in decisions_by_col for col in force_hide_columns):
            self.hide_columns_permanently(ws, force_hide_columns)
    '''
    
    print(integration_code)


def main():
    """Test de hide_columns functionaliteit."""
    print("🧪 ENHANCED COLUMN HIDING TEST")
    print("=" * 40)
    
    # Zoek een test template
    test_files = [
        "templates/template_besteleenheid.xlsx",
        "out/ghx_template_besteleenheid_20250816_170414.xlsx",
        "templates/GHXstandaardTemplate v24.07A.xlsx"
    ]
    
    input_file = None
    for file_path in test_files:
        if Path(file_path).exists():
            input_file = Path(file_path)
            break
    
    if not input_file:
        print("❌ Geen test bestand gevonden")
        return
    
    print(f"📁 Test bestand: {input_file}")
    
    # Maak ColumnHider instance
    hider = ColumnHider()
    
    # Test alle methodes
    results = hider.test_all_methods(
        input_file=input_file,
        columns_to_hide=['AA', 'AB'],
        sheet_name="Template NL"
    )
    
    # Toon resultaten
    print("\n📊 TEST RESULTATEN:")
    print("-" * 30)
    
    for method_name, result in results['methods_tested'].items():
        if 'error' in result:
            print(f"❌ {method_name}: {result['error']}")
        else:
            success_rate = result['success_count'] / len(results['columns_tested']) * 100
            print(f"✅ {method_name}: {success_rate:.0f}% succes ({result['success_count']}/{len(results['columns_tested'])})")
            if 'output_file' in result:
                print(f"   📄 Output: {result['output_file']}")
    
    # Toon aanbevelingen
    print("\n💡 AANBEVELINGEN:")
    print("-" * 20)
    for rec in results['recommendations']:
        print(f"  {rec}")
    
    # Sla gedetailleerde resultaten op
    results_file = Path("out/column_hiding_test_results.json")
    results_file.parent.mkdir(exist_ok=True)
    
    with open(results_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    
    print(f"\n💾 Gedetailleerde resultaten: {results_file}")
    
    # Toon integratie suggestie
    print("\n🔧 INTEGRATIE:")
    integrate_with_excel_engine()


if __name__ == "__main__":
    main()