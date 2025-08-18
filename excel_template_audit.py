#!/usr/bin/env python3
"""
Excel Template Audit Script - Onderzoek waarom kolommen AA/AB automatisch zichtbaar worden.

Dit script analyseert Excel templates om te ontdekken welke structuren 
Excel dwingen om verborgen kolommen weer zichtbaar te maken.
"""

from pathlib import Path
from typing import List, Dict, Any, Set, Optional
import json
from dataclasses import dataclass, asdict
import re
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook


@dataclass 
class ColumnConflict:
    """Gedetecteerd conflict met kolom verbergen."""
    conflict_type: str
    sheet_name: str
    details: str
    severity: str  # 'HIGH', 'MEDIUM', 'LOW'
    recommendation: str
    technical_info: Dict[str, Any]


class ExcelTemplateAuditor:
    """
    Audit Excel templates voor kolom-verberg conflicten.
    """
    
    def __init__(self, target_columns: List[str] = ['AA', 'AB']):
        """
        Initialiseer auditor.
        
        Args:
            target_columns: Kolommen die verborgen moeten worden
        """
        self.target_columns = target_columns
        self.target_indices = [self._col_letter_to_index(col) for col in target_columns]
        self.conflicts = []
        self.workbook = None
        self.file_path = None
        
    def audit_file(self, file_path: Path) -> Dict[str, Any]:
        """
        Voer volledige audit uit op Excel bestand.
        
        Args:
            file_path: Pad naar Excel bestand
            
        Returns:
            Audit rapport dictionary
        """
        self.file_path = file_path
        self.conflicts = []
        
        print(f"🔍 EXCEL TEMPLATE AUDIT: {file_path.name}")
        print("=" * 60)
        
        try:
            self.workbook = load_workbook(file_path, data_only=False)
            
            # Voer alle audit checks uit
            self._audit_structured_tables()
            self._audit_data_validation()
            self._audit_named_ranges()
            self._audit_xml_column_definitions()
            self._audit_merged_cells()
            self._audit_conditional_formatting()
            self._audit_print_areas()
            
            # Genereer rapport
            report = self._generate_report()
            
            return report
            
        except Exception as e:
            print(f"❌ Fout bij audit: {e}")
            return {
                'file': str(file_path),
                'error': str(e),
                'conflicts': [],
                'summary': {'total_conflicts': 0, 'high_severity': 0}
            }
    
    def _audit_structured_tables(self) -> None:
        """Check voor gestructureerde tabellen die over doelkolommen lopen."""
        print("\n📊 AUDIT: Gestructureerde Tabellen (ListObjects)")
        print("-" * 45)
        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            
            if not hasattr(ws, '_tables') or not ws._tables:
                print(f"  {sheet_name}: Geen tabellen gevonden")
                continue
            
            for table_name, table in ws._tables.items():
                # Bepaal tabel bereik - table kan string of object zijn
                if hasattr(table, 'ref'):
                    table_ref = table.ref
                else:
                    table_ref = str(table)
                    
                start_col, start_row, end_col, end_row = self._parse_range(table_ref)
                
                # Check overlap met doelkolommen
                overlapping_cols = []
                for target_idx in self.target_indices:
                    if start_col <= target_idx <= end_col:
                        overlapping_cols.append(self._col_index_to_letter(target_idx))
                
                if overlapping_cols:
                    conflict = ColumnConflict(
                        conflict_type="STRUCTURED_TABLE",
                        sheet_name=sheet_name,
                        details=f"Tabel '{table_name}' loopt over kolommen {overlapping_cols}",
                        severity="HIGH",
                        recommendation="Verplaats tabel of sluit kolommen uit van tabel bereik",
                        technical_info={
                            'table_name': table_name,
                            'table_range': table_ref,
                            'overlapping_columns': overlapping_cols
                        }
                    )
                    self.conflicts.append(conflict)
                    print(f"  ❌ {sheet_name}: Tabel '{table_name}' raakt {overlapping_cols} (bereik: {table_ref})")
                else:
                    print(f"  ✅ {sheet_name}: Tabel '{table_name}' OK (bereik: {table_ref})")
    
    def _audit_data_validation(self) -> None:
        """Check voor data validatie regels die doelkolommen gebruiken."""
        print("\n✅ AUDIT: Data Validatie")
        print("-" * 25)
        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            
            validation_conflicts = []
            
            # Check elke cel voor validatie regels
            if hasattr(ws, 'data_validations') and ws.data_validations:
                for dv in ws.data_validations.dataValidation:
                    if hasattr(dv, 'formula1') and dv.formula1:
                        # Check of formula verwijst naar doelkolommen
                        formula = str(dv.formula1)
                        
                        for target_col in self.target_columns:
                            if target_col in formula:
                                affected_ranges = [str(sqref) for sqref in dv.sqref.ranges]
                                
                                conflict = ColumnConflict(
                                    conflict_type="DATA_VALIDATION",
                                    sheet_name=sheet_name,
                                    details=f"Validatie gebruikt kolom {target_col} in formule: {formula}",
                                    severity="MEDIUM",
                                    recommendation=f"Verplaats bron data van {target_col} naar verborgen sheet",
                                    technical_info={
                                        'formula': formula,
                                        'affected_ranges': affected_ranges,
                                        'target_column': target_col
                                    }
                                )
                                self.conflicts.append(conflict)
                                validation_conflicts.append(f"Kolom {target_col} in formule: {formula}")
            
            if validation_conflicts:
                print(f"  ❌ {sheet_name}: {len(validation_conflicts)} validatie conflicten")
                for conflict in validation_conflicts:
                    print(f"     - {conflict}")
            else:
                print(f"  ✅ {sheet_name}: Geen validatie conflicten")
    
    def _audit_named_ranges(self) -> None:
        """Check voor named ranges die naar doelkolommen verwijzen."""
        print("\n📝 AUDIT: Named Ranges")
        print("-" * 22)
        
        if not hasattr(self.workbook, 'defined_names') or not self.workbook.defined_names:
            print("  ✅ Geen named ranges gevonden")
            return
        
        named_range_conflicts = []
        
        for defined_name in self.workbook.defined_names.definedName:
            name = defined_name.name
            destinations = defined_name.destinations
            
            for sheet_name, coord_range in destinations:
                # Check of range doelkolommen bevat
                if self._range_contains_target_columns(coord_range):
                    affected_cols = self._get_columns_in_range(coord_range)
                    overlapping = [col for col in affected_cols if col in self.target_columns]
                    
                    conflict = ColumnConflict(
                        conflict_type="NAMED_RANGE",
                        sheet_name=sheet_name or "Workbook",
                        details=f"Named range '{name}' verwijst naar kolommen {overlapping}",
                        severity="MEDIUM",
                        recommendation=f"Herdefinieer named range om kolommen {overlapping} uit te sluiten",
                        technical_info={
                            'range_name': name,
                            'range_reference': coord_range,
                            'overlapping_columns': overlapping
                        }
                    )
                    self.conflicts.append(conflict)
                    named_range_conflicts.append(f"'{name}' → {coord_range} raakt {overlapping}")
        
        if named_range_conflicts:
            print(f"  ❌ {len(named_range_conflicts)} named range conflicten:")
            for conflict in named_range_conflicts:
                print(f"     - {conflict}")
        else:
            print("  ✅ Geen named range conflicten")
    
    def _audit_xml_column_definitions(self) -> None:
        """Analyseer XML kolom definities in sheet XML."""
        print("\n🔧 AUDIT: XML Kolom Definities")
        print("-" * 32)
        
        try:
            # Open Excel als ZIP voor directe XML toegang
            with ZipFile(self.file_path, 'r') as zip_file:
                
                # Find alle sheet XML bestanden
                sheet_xmls = [name for name in zip_file.namelist() 
                            if name.startswith('xl/worksheets/') and name.endswith('.xml')]
                
                for sheet_xml in sheet_xmls:
                    sheet_name = sheet_xml.split('/')[-1].replace('.xml', '')
                    
                    try:
                        xml_content = zip_file.read(sheet_xml).decode('utf-8')
                        root = ET.fromstring(xml_content)
                        
                        # Find <cols> sectie
                        cols_element = root.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}cols')
                        
                        if cols_element is not None:
                            col_conflicts = self._analyze_cols_xml(cols_element, sheet_name)
                            
                            if col_conflicts:
                                print(f"  ❌ {sheet_name}: {len(col_conflicts)} XML kolom conflicten")
                                for conflict in col_conflicts:
                                    print(f"     - {conflict}")
                            else:
                                print(f"  ✅ {sheet_name}: XML kolom definities OK")
                        else:
                            print(f"  ✅ {sheet_name}: Geen <cols> definities")
                            
                    except Exception as e:
                        print(f"  ⚠️ {sheet_name}: Fout bij XML analyse: {e}")
                        
        except Exception as e:
            print(f"  ❌ Fout bij XML audit: {e}")
    
    def _analyze_cols_xml(self, cols_element: ET.Element, sheet_name: str) -> List[str]:
        """Analyseer <cols> XML element voor kolom conflicten."""
        conflicts = []
        
        for col_elem in cols_element.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}col'):
            min_col = int(col_elem.get('min', 0))
            max_col = int(col_elem.get('max', 0))
            hidden = col_elem.get('hidden', 'false').lower() == 'true'
            width = col_elem.get('width', 'auto')
            
            # Check overlap met doelkolommen
            overlapping_indices = []
            for target_idx in self.target_indices:
                if min_col <= target_idx <= max_col:
                    overlapping_indices.append(target_idx)
            
            if overlapping_indices:
                overlapping_letters = [self._col_index_to_letter(idx) for idx in overlapping_indices]
                
                conflict_detail = f"XML col definitie min={min_col} max={max_col} (kolommen {overlapping_letters})"
                if not hidden:
                    conflict_detail += " - NIET verborgen in XML"
                if width and width != '0':
                    conflict_detail += f" - breedte: {width}"
                
                conflict = ColumnConflict(
                    conflict_type="XML_COLUMN_DEFINITION",
                    sheet_name=sheet_name,
                    details=conflict_detail,
                    severity="HIGH" if not hidden else "MEDIUM",
                    recommendation="XML kolom definitie moet hidden='true' en width='0' hebben",
                    technical_info={
                        'min_col': min_col,
                        'max_col': max_col,
                        'hidden': hidden,
                        'width': width,
                        'overlapping_columns': overlapping_letters
                    }
                )
                self.conflicts.append(conflict)
                conflicts.append(conflict_detail)
        
        return conflicts
    
    def _audit_merged_cells(self) -> None:
        """Check voor samengevoegde cellen die doelkolommen bevatten."""
        print("\n🔗 AUDIT: Samengevoegde Cellen")
        print("-" * 28)
        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            merged_conflicts = []
            
            for merged_range in ws.merged_cells.ranges:
                range_str = str(merged_range)
                if self._range_contains_target_columns(range_str):
                    overlapping = self._get_columns_in_range(range_str)
                    target_overlap = [col for col in overlapping if col in self.target_columns]
                    
                    conflict = ColumnConflict(
                        conflict_type="MERGED_CELLS",
                        sheet_name=sheet_name,
                        details=f"Samengevoegde cellen {range_str} bevatten kolommen {target_overlap}",
                        severity="MEDIUM",
                        recommendation="Split samengevoegde cellen of verplaats naar andere kolommen",
                        technical_info={
                            'merged_range': range_str,
                            'overlapping_columns': target_overlap
                        }
                    )
                    self.conflicts.append(conflict)
                    merged_conflicts.append(f"{range_str} → {target_overlap}")
            
            if merged_conflicts:
                print(f"  ❌ {sheet_name}: {len(merged_conflicts)} samengevoegde cel conflicten")
                for conflict in merged_conflicts:
                    print(f"     - {conflict}")
            else:
                print(f"  ✅ {sheet_name}: Geen samengevoegde cel conflicten")
    
    def _audit_conditional_formatting(self) -> None:
        """Check voor voorwaardelijke opmaak die doelkolommen gebruikt."""
        print("\n🎨 AUDIT: Voorwaardelijke Opmaak")
        print("-" * 30)
        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            formatting_conflicts = []
            
            if hasattr(ws, 'conditional_formatting') and ws.conditional_formatting:
                for cf in ws.conditional_formatting:
                    # Check bereiken waar opmaak op wordt toegepast
                    for range_obj in cf.sqref.ranges:
                        range_str = str(range_obj)
                        if self._range_contains_target_columns(range_str):
                            overlapping = self._get_columns_in_range(range_str)
                            target_overlap = [col for col in overlapping if col in self.target_columns]
                            
                            conflict = ColumnConflict(
                                conflict_type="CONDITIONAL_FORMATTING",
                                sheet_name=sheet_name,
                                details=f"Voorwaardelijke opmaak op bereik {range_str} bevat kolommen {target_overlap}",
                                severity="LOW",
                                recommendation="Verwijder voorwaardelijke opmaak van verborgen kolommen",
                                technical_info={
                                    'formatting_range': range_str,
                                    'overlapping_columns': target_overlap
                                }
                            )
                            self.conflicts.append(conflict)
                            formatting_conflicts.append(f"{range_str} → {target_overlap}")
            
            if formatting_conflicts:
                print(f"  ❌ {sheet_name}: {len(formatting_conflicts)} opmaak conflicten")
                for conflict in formatting_conflicts:
                    print(f"     - {conflict}")
            else:
                print(f"  ✅ {sheet_name}: Geen opmaak conflicten")
    
    def _audit_print_areas(self) -> None:
        """Check voor print gebieden die doelkolommen bevatten."""
        print("\n🖨️ AUDIT: Print Gebieden")
        print("-" * 23)
        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            
            if ws.print_area:
                if self._range_contains_target_columns(ws.print_area):
                    overlapping = self._get_columns_in_range(ws.print_area)
                    target_overlap = [col for col in overlapping if col in self.target_columns]
                    
                    conflict = ColumnConflict(
                        conflict_type="PRINT_AREA",
                        sheet_name=sheet_name,
                        details=f"Print gebied {ws.print_area} bevat kolommen {target_overlap}",
                        severity="LOW",
                        recommendation="Pas print gebied aan om verborgen kolommen uit te sluiten",
                        technical_info={
                            'print_area': ws.print_area,
                            'overlapping_columns': target_overlap
                        }
                    )
                    self.conflicts.append(conflict)
                    print(f"  ❌ {sheet_name}: Print gebied raakt {target_overlap} ({ws.print_area})")
                else:
                    print(f"  ✅ {sheet_name}: Print gebied OK")
            else:
                print(f"  ✅ {sheet_name}: Geen print gebied ingesteld")
    
    def _generate_report(self) -> Dict[str, Any]:
        """Genereer uitgebreid audit rapport."""
        print(f"\n📋 AUDIT RAPPORT")
        print("=" * 20)
        
        # Groepeer conflicten per type
        conflicts_by_type = {}
        conflicts_by_severity = {'HIGH': 0, 'MEDIUM': 0, 'LOW': 0}
        
        for conflict in self.conflicts:
            conflict_type = conflict.conflict_type
            if conflict_type not in conflicts_by_type:
                conflicts_by_type[conflict_type] = []
            conflicts_by_type[conflict_type].append(conflict)
            conflicts_by_severity[conflict.severity] += 1
        
        # Print samenvatting
        total_conflicts = len(self.conflicts)
        print(f"Totaal conflicten: {total_conflicts}")
        print(f"Hoge prioriteit: {conflicts_by_severity['HIGH']}")
        print(f"Gemiddelde prioriteit: {conflicts_by_severity['MEDIUM']}")
        print(f"Lage prioriteit: {conflicts_by_severity['LOW']}")
        
        if total_conflicts == 0:
            print("✅ Geen conflicten gevonden - kolommen zouden verborgen moeten blijven!")
        else:
            print(f"\n❌ {total_conflicts} conflicten gevonden die Excel kunnen dwingen kolommen zichtbaar te maken")
        
        # Genereer aanbevelingen
        recommendations = self._generate_recommendations(conflicts_by_type)
        
        report = {
            'file': str(self.file_path),
            'target_columns': self.target_columns,
            'total_conflicts': total_conflicts,
            'conflicts_by_severity': conflicts_by_severity,
            'conflicts_by_type': {k: len(v) for k, v in conflicts_by_type.items()},
            'conflicts': [asdict(conflict) for conflict in self.conflicts],
            'recommendations': recommendations,
            'audit_timestamp': str(Path.cwd())
        }
        
        return report
    
    def _generate_recommendations(self, conflicts_by_type: Dict[str, List[ColumnConflict]]) -> List[str]:
        """Genereer specifieke aanbevelingen gebaseerd op gevonden conflicten."""
        recommendations = []
        
        if 'STRUCTURED_TABLE' in conflicts_by_type:
            recommendations.append("🔥 KRITIEK: Verplaats gestructureerde tabellen weg van AA/AB of sluit deze kolommen uit")
        
        if 'XML_COLUMN_DEFINITION' in conflicts_by_type:
            recommendations.append("⚠️ XML kolom definities moeten aangepast worden - dit vereist directe XML bewerking")
        
        if 'DATA_VALIDATION' in conflicts_by_type:
            recommendations.append("📋 Verplaats validatie brondata naar verborgen 'Lookup' sheet")
        
        if 'NAMED_RANGE' in conflicts_by_type:
            recommendations.append("📝 Herdefinieer named ranges om AA/AB uit te sluiten")
        
        if len(self.conflicts) == 0:
            recommendations.append("✅ Geen structurele conflicten - probleem ligt mogelijk bij Excel versie of bestandsformaat")
        
        return recommendations
    
    # Helper methods
    
    def _col_letter_to_index(self, col_letter: str) -> int:
        """Convert kolom letter naar 1-based index."""
        result = 0
        for char in col_letter:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result
    
    def _col_index_to_letter(self, index: int) -> str:
        """Convert 1-based index naar kolom letter."""
        result = ""
        while index > 0:
            index -= 1
            result = chr(index % 26 + ord('A')) + result
            index //= 26
        return result
    
    def _parse_range(self, range_str: str) -> tuple:
        """Parse Excel range naar (start_col, start_row, end_col, end_row)."""
        # Simplified parser voor basis ranges
        if ':' in range_str:
            start, end = range_str.split(':')
        else:
            start = end = range_str
        
        start_col = self._extract_col_from_cell(start)
        start_row = self._extract_row_from_cell(start)
        end_col = self._extract_col_from_cell(end)
        end_row = self._extract_row_from_cell(end)
        
        return start_col, start_row, end_col, end_row
    
    def _extract_col_from_cell(self, cell_ref: str) -> int:
        """Extraheer kolom index van cel referentie."""
        col_letters = ''.join([c for c in cell_ref if c.isalpha()])
        return self._col_letter_to_index(col_letters)
    
    def _extract_row_from_cell(self, cell_ref: str) -> int:
        """Extraheer rij nummer van cel referentie."""
        row_digits = ''.join([c for c in cell_ref if c.isdigit()])
        return int(row_digits) if row_digits else 1
    
    def _range_contains_target_columns(self, range_str: str) -> bool:
        """Check of range overlap heeft met doelkolommen."""
        try:
            start_col, _, end_col, _ = self._parse_range(range_str)
            return any(start_col <= target_idx <= end_col for target_idx in self.target_indices)
        except:
            # Fallback: check of doelkolom letters in range string staan
            return any(col in range_str for col in self.target_columns)
    
    def _get_columns_in_range(self, range_str: str) -> List[str]:
        """Krijg alle kolom letters in een range."""
        try:
            start_col, _, end_col, _ = self._parse_range(range_str)
            return [self._col_index_to_letter(i) for i in range(start_col, end_col + 1)]
        except:
            return []


def main():
    """Test de audit functionaliteit op beide bestanden."""
    print("🔍 EXCEL TEMPLATE AUDIT SCRIPT")
    print("=" * 35)
    
    # Test bestanden
    test_files = [
        "out/test_direct_hiding.xlsx",        # Problematisch bestand
        "out/compat_test_outline_hidden.xlsx" # Werkend bestand
    ]
    
    auditor = ExcelTemplateAuditor(['AA', 'AB'])
    
    for test_file in test_files:
        file_path = Path(test_file)
        if not file_path.exists():
            print(f"⚠️ Bestand niet gevonden: {file_path}")
            continue
        
        print(f"\n{'='*80}")
        report = auditor.audit_file(file_path)
        
        # Sla rapport op
        report_file = file_path.parent / f"audit_report_{file_path.stem}.json"
        with open(report_file, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, ensure_ascii=False)
        
        print(f"\n💾 Rapport opgeslagen: {report_file}")
    
    print(f"\n🎯 CONCLUSIE:")
    print("Vergelijk de rapporten om te zien waarom het ene bestand werkt en het andere niet!")


if __name__ == "__main__":
    main()