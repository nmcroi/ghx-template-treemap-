#!/usr/bin/env python3
"""
CLI entrypoint voor GHX Template Generator.

Gebruik:
    python -m src.main --context sample.json --mapping config/field_mapping.json --out output.xlsx
"""

import argparse
import sys
from pathlib import Path
from typing import Optional
import json

from .context import Context
from .mapping import FieldMapping
from .engine import TemplateEngine
from .excel import ExcelProcessor
from .stamp import TemplateStamp


def setup_argparser() -> argparse.ArgumentParser:
    """Setup command line argument parser."""
    parser = argparse.ArgumentParser(
        description="GHX Template Generator - Genereer aangepaste Excel templates",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Voorbeelden:
    # Basis gebruik
    python -m src.main --context sample.json --out template.xlsx
    
    # Met custom mapping en template locatie
    python -m src.main --context sample.json --mapping custom_mapping.json --templates ./templates --out result.xlsx
    
    # Met custom kleuren
    python -m src.main --context sample.json --mandatory-color "#FFE6CC" --out template.xlsx
    
    # Info over bestaand bestand
    python -m src.main --info existing_template.xlsx
        """
    )
    
    # Input/output
    parser.add_argument(
        "--context", "-c",
        type=Path,
        help="Pad naar context JSON bestand"
    )
    
    parser.add_argument(
        "--mapping", "-m", 
        type=Path,
        default=Path("config/field_mapping.json"),
        help="Pad naar field mapping JSON (default: config/field_mapping.json)"
    )
    
    parser.add_argument(
        "--templates", "-t",
        type=Path, 
        default=Path("templates"),
        help="Pad naar templates map (default: templates/)"
    )
    
    parser.add_argument(
        "--out", "-o",
        type=Path,
        help="Output Excel bestand pad"
    )
    
    # Template keuze
    parser.add_argument(
        "--prefer",
        choices=["bestel", "verpakking", "staffel", "auto"],
        default="auto",
        help="Template voorkeur (default: auto via context)"
    )
    
    # Styling
    parser.add_argument(
        "--mandatory-color",
        default="#FFF2CC",
        help="Hex kleur voor verplichte velden (default: #FFF2CC)"
    )
    
    parser.add_argument(
        "--hidden-color", 
        default="#EEEEEE",
        help="Hex kleur voor verborgen velden (default: #EEEEEE)"
    )
    
    # Utilities
    parser.add_argument(
        "--info", "-i",
        type=Path,
        help="Toon informatie over bestaand template bestand"
    )
    
    parser.add_argument(
        "--validate-context",
        type=Path,
        help="Valideer context JSON bestand"
    )
    
    parser.add_argument(
        "--validate-mapping",
        type=Path, 
        help="Valideer field mapping JSON bestand"
    )
    
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Verbose output"
    )
    
    return parser


def load_context(context_path: Path, verbose: bool = False) -> Context:
    """Laad en valideer context."""
    if verbose:
        print(f"Context laden van: {context_path}")
    
    if not context_path.exists():
        raise FileNotFoundError(f"Context bestand niet gevonden: {context_path}")
    
    context = Context.from_json_file(context_path)
    
    if verbose:
        print(f"Context geladen: {context.get_preset_code()}")
        print(f"Labels: {sorted(context.labels())}")
    
    return context


def load_mapping(mapping_path: Path, verbose: bool = False) -> FieldMapping:
    """Laad en valideer field mapping."""
    if verbose:
        print(f"Mapping laden van: {mapping_path}")
    
    if not mapping_path.exists():
        raise FileNotFoundError(f"Mapping bestand niet gevonden: {mapping_path}")
    
    mapping = FieldMapping.from_file(mapping_path)
    
    # Valideer dependencies
    dep_errors = mapping.validate_dependencies()
    if dep_errors:
        raise ValueError(f"Dependency validatie gefaald:\n" + "\n".join(dep_errors))
    
    if verbose:
        field_count = len(mapping.get_all_fields())
        column_count = len(mapping.get_columns())
        print(f"Mapping geladen: {field_count} velden, {column_count} kolommen")
    
    return mapping


def find_template_file(templates_dir: Path, context: Context, prefer: str) -> Path:
    """Vind juiste template bestand."""
    # Bepaal template basename
    if prefer == "auto":
        basename = context.get_template_basename()
    elif prefer == "bestel":
        basename = "template_besteleenheid" 
    elif prefer == "verpakking":
        basename = "template_verpakkingseenheid"
    elif prefer == "staffel":
        basename = "template_staffel"
    else:
        raise ValueError(f"Onbekende template voorkeur: {prefer}")
    
    # Zoek bestand met .xlsx extensie
    template_path = templates_dir / f"{basename}.xlsx"
    
    if not template_path.exists():
        # Fallback: zoek in bestaande templates
        available = list(templates_dir.glob("*.xlsx"))
        if available:
            print(f"Waarschuwing: {template_path} niet gevonden, gebruik {available[0]}")
            return available[0]
        else:
            raise FileNotFoundError(f"Geen template bestanden gevonden in {templates_dir}")
    
    return template_path


def process_template(context: Context,
                    mapping: FieldMapping, 
                    template_path: Path,
                    output_path: Path,
                    mandatory_color: str,
                    hidden_color: str,
                    verbose: bool = False) -> None:
    """Verwerk template en genereer output."""
    
    if verbose:
        print(f"Template verwerken: {template_path} -> {output_path}")
    
    # Maak engine
    engine = TemplateEngine(context, mapping)
    
    # Verwerk alle velden
    decisions = engine.process_all_fields()
    
    if verbose:
        visible_count = sum(1 for d in decisions if d.visible)
        mandatory_count = sum(1 for d in decisions if d.visible and d.mandatory)
        print(f"Beslissingen: {len(decisions)} velden, {visible_count} zichtbaar, {mandatory_count} verplicht")
    
    # Maak Excel processor
    excel_processor = ExcelProcessor(mandatory_color, hidden_color)
    
    # Valideer template
    template_errors = excel_processor.validate_template(template_path)
    if template_errors:
        print("Waarschuwingen template validatie:")
        for error in template_errors:
            print(f"  - {error}")
    
    # Verwerk template
    context_dict = context.to_dict()
    context_dict["_labels"] = list(context.labels())  # Voeg labels toe voor debugging
    
    # Bepaal sheet naam (GHX templates gebruiken "Template NL")
    sheet_name = "Template NL"  # Default voor GHX templates
    
    excel_processor.process_template(
        template_path,
        output_path, 
        decisions,
        context_dict,
        sheet_name
    )
    
    if verbose:
        print(f"Template succesvol gegenereerd: {output_path}")


def show_template_info(file_path: Path) -> None:
    """Toon informatie over bestaand template."""
    print(f"\nTemplate Informatie: {file_path}")
    print("=" * 50)
    
    if not file_path.exists():
        print("❌ Bestand niet gevonden")
        return
    
    # Probeer stempel te extraheren
    stamp_info = TemplateStamp.get_stamp_info(file_path)
    
    if not stamp_info:
        print("❌ Geen GHX stempel gevonden - mogelijk handmatig gemaakt bestand")
        return
    
    print(f"✅ GHX Template gevonden")
    print(f"📄 Preset Code: {stamp_info.get('preset_code', 'Onbekend')}")
    print(f"✔️  Geldig: {'Ja' if stamp_info.get('valid') else 'Nee'}")
    
    if not stamp_info.get('valid') and 'errors' in stamp_info:
        print("❌ Validatie fouten:")
        for error in stamp_info['errors']:
            print(f"   - {error}")
    
    summary = stamp_info.get('summary', {})
    if summary:
        print(f"\n📋 Configuratie:")
        print(f"   Template Keuze: {summary.get('template_choice')}")
        print(f"   GS1 Modus: {summary.get('gs1_mode')}")
        print(f"   Product Type: {summary.get('product_type')}")
        print(f"   Instellingen: {', '.join(summary.get('institutions', []))}")
        print(f"   Versie: {summary.get('version')}")
    
    # Probeer Excel info te tonen
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        print(f"\n📊 Excel Info:")
        print(f"   Sheets: {', '.join(wb.sheetnames)}")
        
        if "Sheet1" in wb.sheetnames:
            ws = wb["Sheet1"]
            print(f"   Sheet1 Rijen: {ws.max_row}")
            print(f"   Sheet1 Kolommen: {ws.max_column}")
        
        wb.close()
        
    except Exception as e:
        print(f"⚠️  Kon Excel info niet lezen: {e}")


def validate_context_file(context_path: Path) -> None:
    """Valideer context JSON bestand."""
    print(f"\nContext Validatie: {context_path}")
    print("=" * 40)
    
    try:
        context = load_context(context_path, verbose=True)
        print("✅ Context is geldig")
        
        print(f"\n📋 Context Details:")
        print(f"   Preset Code: {context.get_preset_code()}")
        print(f"   Template: {context.get_template_basename()}")
        print(f"   Labels: {', '.join(sorted(context.labels()))}")
        
    except Exception as e:
        print(f"❌ Context validatie gefaald: {e}")


def validate_mapping_file(mapping_path: Path) -> None:
    """Valideer mapping JSON bestand."""
    print(f"\nMapping Validatie: {mapping_path}")
    print("=" * 40)
    
    try:
        mapping = load_mapping(mapping_path, verbose=True)
        print("✅ Mapping is geldig")
        
        fields = mapping.get_all_fields()
        columns = mapping.get_columns()
        
        print(f"\n📋 Mapping Details:")
        print(f"   Velden: {len(fields)}")
        print(f"   Kolommen: {', '.join(sorted(columns))}")
        
        # Toon velden per categorie
        visible_always = [name for name, config in fields.items() if config.get("visible") == "always"]
        visible_conditional = [name for name, config in fields.items() if "visible_only" in config or "visible_except" in config]
        mandatory_always = [name for name, config in fields.items() if config.get("mandatory") == "always"]
        
        if visible_always:
            print(f"   Altijd zichtbaar: {len(visible_always)} velden")
        if visible_conditional:
            print(f"   Voorwaardelijk zichtbaar: {len(visible_conditional)} velden")
        if mandatory_always:
            print(f"   Altijd verplicht: {len(mandatory_always)} velden")
        
    except Exception as e:
        print(f"❌ Mapping validatie gefaald: {e}")


def main():
    """Main CLI functie."""
    parser = setup_argparser()
    args = parser.parse_args()
    
    try:
        # Utility commands
        if args.info:
            show_template_info(args.info)
            return
        
        if args.validate_context:
            validate_context_file(args.validate_context)
            return
            
        if args.validate_mapping:
            validate_mapping_file(args.validate_mapping)
            return
        
        # Template generatie
        if not args.context:
            parser.error("--context is verplicht voor template generatie")
        
        if not args.out:
            parser.error("--out is verplicht voor template generatie")
        
        # Laad context en mapping
        context = load_context(args.context, args.verbose)
        mapping = load_mapping(args.mapping, args.verbose)
        
        # Vind template
        template_path = find_template_file(args.templates, context, args.prefer)
        
        if args.verbose:
            print(f"Gekozen template: {template_path}")
        
        # Maak output directory als het niet bestaat
        args.out.parent.mkdir(parents=True, exist_ok=True)
        
        # Verwerk template
        process_template(
            context,
            mapping,
            template_path,
            args.out,
            args.mandatory_color,
            args.hidden_color,
            args.verbose
        )
        
        print(f"✅ Template succesvol gegenereerd: {args.out}")
        
    except KeyboardInterrupt:
        print("\n❌ Onderbroken door gebruiker")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Fout: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
