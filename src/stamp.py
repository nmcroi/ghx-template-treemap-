"""
Stempel module voor GHX Template Generator.

Bevat functionaliteit voor het embedden en extraheren van metadata in Excel bestanden.
"""

import json
from pathlib import Path
from typing import Dict, Any, Optional, Tuple, List
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.workbook.defined_name import DefinedName


class TemplateStamp:
    """
    Template stempel manager voor metadata tracking.
    """
    
    METADATA_SHEET_NAME = "_GHX_META"
    STAMP_NAMED_RANGE = "GHX_STAMP"
    GENERATOR_VERSION = "1.0.0"
    
    @classmethod
    def embed_stamp(cls, 
                   wb: Workbook, 
                   context_dict: Dict[str, Any],
                   preset_code: str) -> None:
        """
        Embed GHX stempel in workbook.
        
        Args:
            wb: Workbook object
            context_dict: Context dictionary voor JSON embed
            preset_code: Compacte preset code
        """
        # Maak metadata sheet
        cls._create_metadata_sheet(wb, context_dict)
        
        # Maak named range voor preset code
        cls._create_stamp_named_range(wb, preset_code)
    
    @classmethod
    def extract_stamp(cls, file_path: Path) -> Optional[Tuple[Dict[str, Any], str]]:
        """
        Extraheer GHX stempel van Excel bestand.
        
        Args:
            file_path: Pad naar Excel bestand
            
        Returns:
            Tuple van (context_dict, preset_code) of None als geen stempel
        """
        try:
            wb = load_workbook(file_path)
            
            context_dict = None
            preset_code = None
            
            # Probeer metadata sheet te lezen
            if cls.METADATA_SHEET_NAME in wb.sheetnames:
                meta_ws = wb[cls.METADATA_SHEET_NAME]
                json_str = meta_ws["A1"].value
                
                if json_str:
                    try:
                        context_dict = json.loads(json_str)
                    except json.JSONDecodeError:
                        print(f"Waarschuwing: Ongeldige JSON in metadata sheet van {file_path}")
                
                # Probeer preset code te lezen
                preset_code = meta_ws["B1"].value
            
            # Probeer named range te lezen als fallback
            if not preset_code and cls.STAMP_NAMED_RANGE in wb.defined_names:
                try:
                    defn = wb.defined_names[cls.STAMP_NAMED_RANGE]
                    # Parse de reference om de waarde te krijgen
                    if defn.attr_text and "!" in defn.attr_text:
                        sheet_ref, cell_ref = defn.attr_text.split("!")
                        sheet_name = sheet_ref.strip("'\"")
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            preset_code = ws[cell_ref.strip("$")].value
                except Exception as e:
                    print(f"Waarschuwing: Kon named range niet lezen: {e}")
            
            wb.close()
            
            if context_dict or preset_code:
                return (context_dict, preset_code)
            
            return None
            
        except Exception as e:
            print(f"Waarschuwing: Kon stempel niet extraheren van {file_path}: {e}")
            return None
    
    @classmethod
    def _create_metadata_sheet(cls, wb: Workbook, context_dict: Dict[str, Any]) -> None:
        """
        Maak verborgen metadata sheet.
        
        Args:
            wb: Workbook object
            context_dict: Context dictionary
        """
        # Verwijder bestaande metadata sheet als aanwezig
        if cls.METADATA_SHEET_NAME in wb.sheetnames:
            wb.remove(wb[cls.METADATA_SHEET_NAME])
        
        # Maak nieuwe sheet
        meta_ws = wb.create_sheet(cls.METADATA_SHEET_NAME)
        meta_ws.sheet_state = "hidden"
        
        # Voeg metadata toe
        metadata = {
            "context": context_dict,
            "generator": {
                "name": "GHX Template Generator",
                "version": cls.GENERATOR_VERSION,
                "timestamp": datetime.now().isoformat()
            }
        }
        
        # Schrijf prettified JSON naar A1
        json_str = json.dumps(metadata, indent=2, ensure_ascii=False)
        meta_ws["A1"] = json_str
        
        # Voeg extra info toe
        meta_ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        meta_ws["A4"] = f"Generator: GHX Template Generator v{cls.GENERATOR_VERSION}"
        meta_ws["A5"] = f"Context Labels: {', '.join(sorted(context_dict.get('_labels', [])))}" if '_labels' in context_dict else ""
    
    @classmethod
    def _create_stamp_named_range(cls, wb: Workbook, preset_code: str) -> None:
        """
        Maak named range voor compacte stempel code.
        
        Args:
            wb: Workbook object
            preset_code: Compacte preset code
        """
        # Verwijder bestaande named range
        if cls.STAMP_NAMED_RANGE in wb.defined_names:
            del wb.defined_names[cls.STAMP_NAMED_RANGE]
        
        # Voeg preset code toe aan metadata sheet
        if cls.METADATA_SHEET_NAME in wb.sheetnames:
            meta_ws = wb[cls.METADATA_SHEET_NAME]
            meta_ws["B1"] = preset_code
            
            # Maak named range die verwijst naar deze cel
            defn = DefinedName(
                cls.STAMP_NAMED_RANGE, 
                attr_text=f"{cls.METADATA_SHEET_NAME}!$B$1"
            )
            wb.defined_names[cls.STAMP_NAMED_RANGE] = defn
    
    @classmethod
    def validate_stamp(cls, file_path: Path) -> Tuple[bool, List[str]]:
        """
        Valideer integriteit van GHX stempel.
        
        Args:
            file_path: Pad naar Excel bestand
            
        Returns:
            Tuple van (is_valid, errors)
        """
        errors = []
        
        try:
            wb = load_workbook(file_path)
            
            # Controleer metadata sheet
            if cls.METADATA_SHEET_NAME not in wb.sheetnames:
                errors.append("Metadata sheet niet gevonden")
            else:
                meta_ws = wb[cls.METADATA_SHEET_NAME]
                
                # Controleer JSON in A1
                json_str = meta_ws["A1"].value
                if not json_str:
                    errors.append("Geen JSON data in metadata sheet")
                else:
                    try:
                        metadata = json.loads(json_str)
                        
                        # Valideer structuur
                        if "context" not in metadata:
                            errors.append("Context ontbreekt in metadata")
                        
                        if "generator" not in metadata:
                            errors.append("Generator info ontbreekt in metadata")
                        else:
                            gen_info = metadata["generator"]
                            if "version" not in gen_info:
                                errors.append("Generator versie ontbreekt")
                            if "timestamp" not in gen_info:
                                errors.append("Timestamp ontbreekt")
                        
                    except json.JSONDecodeError as e:
                        errors.append(f"Ongeldige JSON in metadata: {e}")
                
                # Controleer preset code in B1
                preset_code = meta_ws["B1"].value
                if not preset_code:
                    errors.append("Preset code ontbreekt in metadata sheet")
            
            # Controleer named range
            if cls.STAMP_NAMED_RANGE not in wb.defined_names:
                errors.append("GHX_STAMP named range niet gevonden")
            
            wb.close()
            
        except Exception as e:
            errors.append(f"Kan bestand niet lezen: {e}")
        
        return (len(errors) == 0, errors)
    
    @classmethod
    def get_stamp_info(cls, file_path: Path) -> Optional[Dict[str, Any]]:
        """
        Krijg leesbare stempel informatie.
        
        Args:
            file_path: Pad naar Excel bestand
            
        Returns:
            Dictionary met stempel informatie of None
        """
        stamp_data = cls.extract_stamp(file_path)
        if not stamp_data:
            return None
        
        context_dict, preset_code = stamp_data
        
        info = {
            "file": str(file_path),
            "preset_code": preset_code,
            "valid": False,
            "context": context_dict
        }
        
        # Controleer validiteit
        is_valid, errors = cls.validate_stamp(file_path)
        info["valid"] = is_valid
        
        if errors:
            info["errors"] = errors
        
        # Voeg samenvatting toe
        if context_dict:
            info["summary"] = {
                "template_choice": context_dict.get("template_choice"),
                "gs1_mode": context_dict.get("gs1_mode"),
                "product_type": context_dict.get("product_type"),
                "institutions": context_dict.get("institutions", []),
                "version": context_dict.get("version")
            }
        
        return info
