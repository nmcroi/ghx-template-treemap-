import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
import json
from pathlib import Path

def customize_template(template_path, output_path, answers):
    """
    Past de GHX template aan op basis van de antwoorden uit de beslisboom,
    zonder de opmaak te wijzigen.
    
    :param template_path: Pad naar de originele template
    :param output_path: Pad waar de aangepaste template moet worden opgeslagen
    :param answers: Dictionary met antwoorden uit de beslisboom
    """
    # Laad de template
    print(f"Template laden: {template_path}")
    workbook = openpyxl.load_workbook(template_path)
    
    # Selecteer de eerste sheet (Template NL)
    sheet = workbook.worksheets[0]
    
    # Bepaal welke kolommen verborgen moeten worden
    columns_to_hide = get_columns_to_hide(answers) # Let op: dit zijn 1-based indexen
    
    # 1. Pas celkleuren aan voor bestelbaarheid
    if not answers.get('all_orderable', False): # Als niet alles bestelbaar is
        print("Niet alles bestelbaar: cellen F1 en G1 aanpassen...")
        light_yellow_fill = PatternFill(start_color="FFFFEE",
                                          end_color="FFFFEE",
                                          fill_type="solid")
        try:
            sheet['F1'].fill = light_yellow_fill
            sheet['G1'].fill = light_yellow_fill
            print("Cellen F1 en G1 aangepast naar lichtgeel.")
        except Exception as e_fill:
             print(f"Fout bij aanpassen celkleur F1/G1: {e_fill}")
    else:
         print("Alles bestelbaar: cellen F1 en G1 blijven ongewijzigd.")
 
    # 2. Sla de 'answers' op in de documenteigenschappen (metadata)
    try:
        print("Opslaan van configuratiecode in documenteigenschappen (subject)...")
        # Converteer answers dict naar JSON string voor opslag
        config_code = json.dumps(answers)
        # Gebruik het 'subject' veld van de properties
        workbook.properties.subject = config_code
        print(f"Code opgeslagen in 'subject': {config_code[:50]}...") # Print eerste 50 tekens ter controle

        # Sla OOK op in GH10 met witte tekst
        print("Opslaan van configuratiecode in GH10 (witte tekst)...")
        white_font = Font(color="FFFFFF") # Witte kleur
        cell_gh10 = sheet['GH10']
        cell_gh10.value = config_code
        cell_gh10.font = white_font
        print(f"Code OOK opgeslagen in GH10: {config_code[:50]}...")
    except Exception as e_metadata:
         print(f"Fout bij opslaan code in metadata: {e_metadata}")

    # Verberg kolommen zonder de breedte te wijzigen
    print(f"Verbergen van {len(columns_to_hide)} kolommen...")
    for col_index in columns_to_hide:
        col_letter = get_column_letter(col_index)
        sheet.column_dimensions[col_letter].hidden = True
        print(f"Kolom {col_letter} verborgen")
    
    # Pas eventueel verplichte velden aan (markeren als mandatory)
    # Dit kan gedaan worden door bijvoorbeeld cellen te markeren met een kleur
    # of door metadata bij te werken - zonder de hoofdopmaak te wijzigen
    
    # Sla de aangepaste template op
    print(f"Template opslaan naar: {output_path}")
    workbook.save(output_path)
    
    return True

def get_columns_to_hide(answers):
    """
    Bepaal welke kolommen verborgen moeten worden op basis van antwoorden
    
    :param answers: Dictionary met antwoorden uit de beslisboom
    :return: List met kolomindexen (1-based) die verborgen moeten worden
    """
    columns_to_hide = []
    
    # 1. Bestelbaarheid
    if answers.get('all_orderable', False):
        # Verberg kolommen gerelateerd aan bestelbaarheid
        # Is BestelbareEenheid (kolom 8)
        columns_to_hide.append(8)
    
    # 2. Product type
    product_type = answers.get('product_type', '')
    
    if product_type == 'Allemaal facilitair':
        # Verberg medische velden
        # GMDN Code (kolom 47)
        # EMDN Code (kolom 48)
        # Medische velden voor CE certificering (kolommen 49-52)
        # MRI compatibiliteit (kolom 93)
        # Implanteerbaar (kolom 83)
        # Sterilisatie (kolommen 90-91)
        columns_to_hide.extend([47, 48, 49, 50, 51, 52, 83, 90, 91, 93])
        
    elif product_type == 'Allemaal medisch':
        # Verberg facilitaire velden
        # UNSPSC Code (kolom 46) - als die vooral voor facilitair is
        # Duurzaamheidsvelden (kolommen 53-55)
        columns_to_hide.extend([46, 53, 54, 55])
        
    elif product_type == 'Allemaal laboratorium':
        # Verberg niet-lab velden
        # Specifieke velden voor laboratoria
        columns_to_hide.extend([83, 90, 91, 93])  # Implanteerbaar, sterilisatie, etc.
    
    # 3. Zorginstellingen
    # Dit is complexer en zou een mapping nodig hebben van welke zorginstelling
    # welke velden vereist. Voor nu laten we dit leeg.
    
    return columns_to_hide

def test_template_customization():
    """Test functie om de template aanpassing te testen"""
    # Test pad (pas aan naar jouw lokale pad)
    template_path = Path("/Users/ghxnielscroiset/Library/CloudStorage/OneDrive-GlobalHealthcareExchange/Documenten/Windsurf/Project TemplateTree app/templates/GHXstandaardTemplate v24.07A.xlsx")
    output_path = Path("/Users/ghxnielscroiset/Library/CloudStorage/OneDrive-GlobalHealthcareExchange/Documenten/Windsurf/Project TemplateTree app/templates/GHXtemplate_custom_test.xlsx")
    
    # Test antwoorden
    test_answers = {
        'all_orderable': False, # <-- Gewijzigd naar False om kleurwijziging te testen
        'organizations': ['Academisch Ziekenhuis Maastricht', 'UMC Utrecht'],  # Geselecteerde zorginstellingen
        'product_type': 'Allemaal facilitair'  # Product type
    }
    
    # Voer test uit
    result = customize_template(template_path, output_path, test_answers)
    
    if result:
        print("✅ Test succesvol: Template is aangepast en opgeslagen.")
    else:
        print("❌ Test mislukt!")

# Als dit script direct wordt uitgevoerd (niet geïmporteerd), voer dan de test uit
if __name__ == "__main__":
    test_template_customization()